import hashlib
import os

import openpyxl
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, HttpResponse

from login.models import *


# Create your views here.
# 管理员首页
def myadmin(request):
    """

    1.教师管理页面
        添加
        删除
        编辑
        ————是否首页显示——————

    2.学生管理页面
        添加
        删除
        编辑
    3.课程管理页面
        添加
        删除
        编辑
    4.管理评价
        添加
        删除
        编辑
    """
    students = Students.objects.filter(is_active=True).count()
    teachers = Teachers.objects.filter(is_active=True).count()
    kecheng = KeCheng.objects.filter(is_active=True).count()
    tiku = TiKu_1.objects.filter(is_active=True).count()
    pingjia = PingJia.objects.filter(is_active=True).count()
    admin = GuanLiYuan.objects.filter(is_active=True).count()

    return render(request, 'myadmin/index.html', {"students": students,
                                                  "teachers": teachers,
                                                  "kecheng": kecheng,
                                                  "tiku": tiku,
                                                  "pingjia": pingjia,
                                                  "admin": admin})


# 管理学生页面
def myadmin_stu(request, pIndex=1):
    stu_list = Students.objects.filter().order_by()  # 班级过滤器
    mywhere = []
    # 获取并判断搜索
    kw = request.GET.get("keyword", None)
    if kw:
        stu_list = stu_list.filter(Q(xuehao__contains=kw) | Q(name__contains=kw))
        mywhere.append('keyword' + kw)

    # 执行分页处理
    pIndex = int(pIndex)
    page = Paginator(stu_list, 10)  # 以每页9条数据分页
    maxpagex = page.num_pages  # 获取最大页数
    # 判断当前页是否越界
    if pIndex > maxpagex:
        pIndex = maxpagex
    if pIndex < 1:
        pIndex = 1

    list2 = page.page(pIndex)  # 获取当前页数据
    plist = page.page_range  # 获取页码表信息
    context = {"stulist": list2, "plist": plist, "pIndex": pIndex, "max_pages": maxpagex, 'mywehere': mywhere}
    return render(request, 'myadmin/myadmin_stu.html', context)


# 添加学生
def stu_add(request):
    if request.method == "GET":
        return render(request, 'myadmin/stu_add.html')
    if request.method == "POST":
        xuehao = request.POST['xuehao']
        name = request.POST['name']
        phone = request.POST['phone']
        banji = request.POST['banji']
        sex = request.POST['sex']

        try:
            stu = Students.objects.filter(xuehao=xuehao)
        except Exception as e:
            print(e)
        if stu:
            msg = '学号已存在！'

            return render(request, 'myadmin/stu_add.html', {"msg": msg})
        else:
            stu_add = Students.objects.create(xuehao=xuehao, name=name, phone=phone, banji=banji, sex=sex)
            msg = '添加学生成功！'
            return render(request, 'myadmin/tishi.html', {"msg": msg})


# 编辑学生信息
def stu_edit(request, xuehao_id):
    try:
        stu = Students.objects.get(xuehao=xuehao_id)
    except Exception as e:
        print(e)
        msg = '没有此学生！！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/stu_edit.html', locals())
    else:

        name = request.POST['name']
        phone = request.POST['phone']
        banji = request.POST['banji']
        sex = request.POST['sex']
        is_active = request.POST['is_active']

        stu.name = name
        stu.phone = phone
        stu.banji = banji
        stu.sex = sex
        stu.is_active = is_active
        stu.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})


# 学生删除
def stu_del(request, xuehao_id):
    try:
        stu = Students.objects.get(xuehao=xuehao_id)
    except Exception as e:
        print(e)
        msg = "没有此学生！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    stu.is_active = False
    stu.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi.html', {"msg": msg})


# 上传学生信息
def stu_upload(request):
    if request.method == "GET":
        return render(request, 'myadmin/stu_upload.html')
    elif request.method == "POST":
        """
        第一步
            1.先读取
            2.查询学号是否已存在
                如果存在添加一个字典
            3.显示能添加的学生数量
                如果存在已存在的学号不能添加并显示学号不能添加学生学号
            4.确认后添加把学生添加到数据库
        """
        xlsx = request.FILES['xlsx']

        file_name, file_extension = os.path.splitext(str(xlsx))
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(xlsx)

            worksheet = wb.active  # 当前活跃的表单
            # print(worksheet)

            # 单元格遍历
            excel_data = list()

            a = 0
            b = 0
            rows = worksheet.iter_rows()
            for row in rows:
                row_data = list()
                if 1 == row[1].row:
                    continue
                xuehao = str(row[1].value)

                try:
                    stu = Students.objects.filter(xuehao=xuehao)
                except Exception as e:
                    print(e)
                if stu:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("no")
                    a += 1
                else:
                    # row_data.append(str(xuehao))
                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("ok")
                    b += 1
                # print(xuehao, row[1].row) #显示行数
                excel_data.append(row_data)
            global global_excel_data
            global_excel_data = excel_data
            msg = "可以添加%s个学生，%s个学号已存在！" % (b, a)
            return render(request, 'myadmin/stu_upload_ok.html', {"excel_data": excel_data, "msg": msg, "a": a})
        else:
            msg = "请你上传Excel文件错误！（格式必须XXXXX.xlsx）"
            return render(request, 'myadmin/tishi.html', {"msg": msg})


# 查看上传的的学生信息
def stu_toupload(request):
    stu_list = list()
    try:
        for row in global_excel_data:
            i = global_excel_data.index(row)
            # if i == 0:
            #     continue
            xuehao = row[1]
            name = row[2]
            xueyuan = row[3]
            banji = row[4]
            sex = row[5]
            if str(sex) == '男':
                sex = 'male'
            elif str(sex) == '女':
                sex = 'female'
            email = row[6]
            phone = row[7]
            stu = Students.objects.filter(xuehao=xuehao)
            if stu:  # 跳出数据库已存在的学生
                stu_list.append(xuehao)
                continue
            # 哈希算法
            m = hashlib.md5()
            m.update(xuehao.encode())
            pswd = m.hexdigest()

            Students.objects.create(xuehao=xuehao,
                                    name=name,
                                    password=pswd,
                                    xueyuan=xueyuan,
                                    banji=banji,
                                    sex=sex,
                                    email=email,
                                    phone=phone)

            # xuehao = row[2].value
            # xuehao = row[2].value
            # xuehao = row[2].value
            # xuehao = row[1].value
        msg = "成功上传学生表！！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    except Exception as e:
        print(e)
    return HttpResponse('请返回重新提交表格!！')


# 教师管理
def myadmin_teachers(request, pIndex):
    stu_list = Teachers.objects.filter().order_by()  # 班级过滤器
    mywhere = []
    # 获取并判断搜索
    kw = request.GET.get("keyword", None)
    if kw:
        stu_list = stu_list.filter(Q(teacher_id__contains=kw) | Q(name__contains=kw))
        mywhere.append('keyword' + kw)

    # 执行分页处理
    pIndex = int(pIndex)
    page = Paginator(stu_list, 10)  # 以每页9条数据分页
    maxpagex = page.num_pages  # 获取最大页数
    # 判断当前页是否越界
    if pIndex > maxpagex:
        pIndex = maxpagex
    if pIndex < 1:
        pIndex = 1

    list2 = page.page(pIndex)  # 获取当前页数据
    plist = page.page_range  # 获取页码表信息
    context = {"stulist": list2, "plist": plist, "pIndex": pIndex, "max_pages": maxpagex, 'mywehere': mywhere}
    return render(request, 'myadmin/teachers/myadmin_teacher.html', context)


# 添加教师
def teachers_add(request):
    if request.method == "GET":
        return render(request, 'myadmin/teachers/teacher_add.html')
    if request.method == "POST":
        teacher_id = request.POST['teacher_id']
        name = request.POST['name']
        phone = request.POST['phone']
        sex = request.POST['sex']

        try:
            stu = Teachers.objects.filter(teacher_id=teacher_id)
        except Exception as e:
            print(e)
        if stu:
            msg = '教工号已存在！'

            return render(request, 'myadmin/teachers/teacher_add.html', {"msg": msg})
        else:
            stu_add = Teachers.objects.create(teacher_id=teacher_id, name=name, phone=phone, sex=sex)
            msg = '添加教师成功！'
            return render(request, 'myadmin/tishi.html', {"msg": msg})


# 教师编辑
def teachers_edit(request, teacher_id):
    try:
        stu = Teachers.objects.get(teacher_id=teacher_id)
    except Exception as e:
        print(e)
        msg = '没有此教师！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/teachers/teacher_edit.html', locals())
    else:

        name = request.POST['name']
        phone = request.POST['phone']

        sex = request.POST['sex']
        is_active = request.POST['is_active']

        stu.name = name
        stu.phone = phone

        stu.sex = sex
        stu.is_active = is_active
        stu.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})


# 删除教师
def teachers_del(request, teacher_id):
    try:
        stu = Teachers.objects.get(teacher_id=teacher_id)
    except Exception as e:
        print(e)
        msg = "没有此教师！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    stu.is_active = False
    stu.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi.html', {"msg": msg})


# 上传教师
def teachers_upload(request):
    if request.method == 'GET':
        return render(request, 'myadmin/teachers/teacher_upload.html')
    elif request.method == 'POST':
        xlsx = request.FILES['xlsx']

        file_name, file_extension = os.path.splitext(str(xlsx))
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(xlsx)

            worksheet = wb.active  # 当前活跃的表单
            # print(worksheet)

            # 单元格遍历
            excel_data = list()

            a = 0
            b = 0
            rows = worksheet.iter_rows()
            for row in rows:
                row_data = list()
                if 1 == row[1].row:
                    continue
                teacher_id = str(row[1].value)

                try:
                    stu = Teachers.objects.filter(teacher_id=teacher_id)
                except Exception as e:
                    print(e)
                if stu:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("no")
                    a += 1
                else:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("ok")
                    b += 1
                # print(xuehao, row[1].row) #显示行数
                excel_data.append(row_data)
            global global_excel_teachers
            global_excel_teachers = excel_data
            msg = "可以添加%s个教师ID，%s个教师ID已存在！" % (b, a)
            return render(request, 'myadmin/teachers/teacher_upload_ok.html',
                          {"excel_data": excel_data, "msg": msg, "a": a})
        else:
            msg = "请你上传Excel文件错误！（格式必须XXXXX.xlsx）"
            return render(request, 'myadmin/tishi.html', {"msg": msg})


def teachers_toupload(request):
    stu_list = list()
    try:
        for row in global_excel_teachers:
            i = global_excel_teachers.index(row)

            teacher_id = row[1]
            name = row[2]

            sex = row[3]
            if str(sex) == '男':
                sex = 'male'
            elif str(sex) == '女':
                sex = 'female'
            email = row[4]
            phone = row[5]

            # 哈希算法
            m = hashlib.md5()
            m.update(teacher_id.encode())
            pswd = m.hexdigest()
            stu = Teachers.objects.filter(teacher_id=teacher_id)
            if stu:  # 跳出数据库已存在的学生
                stu_list.append(teacher_id)
                continue

            Teachers.objects.create(teacher_id=teacher_id,
                                    name=name,
                                    sex=sex,
                                    email=email,
                                    phone=phone,
                                    password=pswd)

        msg = "成功上传教师表！！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    except Exception as e:
        print("teacher_toupload:", e)
    return HttpResponse('请返回重新提交表格!！')


# 课程管理首页
def myadmin_kecheng(request, pIndex):
    stu_list = KeCheng.objects.filter().order_by().values('id', 'kecheng', 'xuehao', 'xuehao__name', 'teacher_id',

                                                          'teacher_id__name', 'is_active')  # 班级过滤器
    mywhere = []
    # 获取并判断搜索
    kw = request.GET.get("keyword", None)
    if kw:
        stu_list = stu_list.filter(Q(id__contains=kw) | Q(kecheng__contains=kw))
        mywhere.append('keyword' + kw)

    # 执行分页处理
    pIndex = int(pIndex)
    page = Paginator(stu_list, 10)  # 以每页9条数据分页
    maxpagex = page.num_pages  # 获取最大页数
    # 判断当前页是否越界
    if pIndex > maxpagex:
        pIndex = maxpagex
    if pIndex < 1:
        pIndex = 1

    list2 = page.page(pIndex)  # 获取当前页数据
    plist = page.page_range  # 获取页码表信息
    context = {"stulist": list2, "plist": plist, "pIndex": pIndex, "max_pages": maxpagex, 'mywehere': mywhere}
    return render(request, 'myadmin/kecheng/myadmin_kecheng.html', context)


def myadmin_kecheng_add(request):
    if request.method == "GET":
        students = Students.objects.filter(is_active=True)
        teachers = Teachers.objects.filter(is_active=True)
        kecheng = KeCheng.objects.filter(is_active=True)

        try:
            id = kecheng.last().id
            return render(request, 'myadmin/kecheng/kecheng_add.html',
                          {"teachers": teachers, "students": students, "id": int(id) + 1})
        except Exception as e:
            print(e)
        id = 0
        return render(request, 'myadmin/kecheng/kecheng_add.html',
                      {"teachers": teachers, "students": students, "id": id})

    if request.method == "POST":
        id = request.POST['id']
        kecheng = request.POST['kecheng']
        xuehao = request.POST['xuehao']
        teacher_id = request.POST['teacher_id']

        try:
            stu = KeCheng.objects.filter(str(id))
        except Exception as e:
            print(e)

        stu_add = KeCheng.objects.create(id=str(id), kecheng=kecheng, xuehao=Students.objects.get(xuehao=xuehao),
                                         teacher_id=Teachers.objects.get(teacher_id=teacher_id))
        msg = '添加课程成功！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})


def myadmin_kecheng_edit(request, kecheng_id):
    try:
        stu = KeCheng.objects.get(id=kecheng_id)
        teachers = Teachers.objects.filter(is_active=True)
        students = Students.objects.filter(is_active=True)

    except Exception as e:
        print(e)
        msg = '没有此课程！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/kecheng/kecheng_edit.html', locals())
    else:

        kecheng = request.POST['kecheng']
        xuehao = request.POST['xuehao']
        teacher_id = request.POST['teacher_id']
        # is_active = request.POST['is_active']

        stu.kecheng = kecheng
        stu.xuehao = Students.objects.get(xuehao=xuehao)
        stu.teacher_id = Teachers.objects.get(teacher_id=teacher_id)
        # stu.is_active = is_active
        stu.save()
        msg = '课程修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})


def myadmin_kecheng_del(request, kecheng_id):
    try:
        stu = KeCheng.objects.get(id=kecheng_id)
    except Exception as e:
        print(e)
        msg = "没有此课程！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    stu.is_active = False
    stu.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi.html', {"msg": msg})


def kecheng_upload(request):
    if request.method == 'GET':
        return render(request, 'myadmin/kecheng/kecheng_upload.html')
    elif request.method == 'POST':
        xlsx = request.FILES['xlsx']

        file_name, file_extension = os.path.splitext(str(xlsx))
        if file_extension == ".xlsx":
            wb = openpyxl.load_workbook(xlsx)

            worksheet = wb.active  # 当前活跃的表单
            # print(worksheet)

            # 单元格遍历
            excel_data = list()

            a = 0
            b = 0
            rows = worksheet.iter_rows()
            for row in rows:
                row_data = list()
                if 1 == row[1].row:
                    continue
                id = str(row[1].value)

                try:
                    stu = KeCheng.objects.filter(id=id)
                except Exception as e:
                    print(e)
                if stu:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("no")
                    a += 1
                else:

                    for cell in row:
                        row_data.append(str(cell.value))
                    row_data.append("ok")
                    b += 1
                # print(xuehao, row[1].row) #显示行数
                excel_data.append(row_data)
            global global_excel_kecheng
            global_excel_kecheng = excel_data
            msg = "可以添加%s个课程，%s个课程已存在！" % (b, a)
            return render(request, 'myadmin/kecheng/kecheng_upload_ok.html',
                          {"excel_data": excel_data, "msg": msg, "a": a})
        else:
            msg = "请你上传Excel文件错误！（格式必须XXXXX.xlsx）"
            return render(request, 'myadmin/tishi.html', {"msg": msg})


def kecheng_toupload(request):
    stu_list = list()
    # return HttpResponse(global_excel_kecheng)
    try:
        for row in global_excel_kecheng:
            i = global_excel_kecheng.index(row)

            id = row[1]
            name = row[2]

            xuehao = row[3]

            teacher_id = row[4]
            # print(id, name, xuehao, teacher_id)

            stu = KeCheng.objects.filter(id=str(id))
            if stu:  # 跳出数据库已存在的课程
                stu_list.append(id)
                continue

            KeCheng.objects.create(id=id,
                                   kecheng=name,
                                   xuehao=Students.objects.get(xuehao=str(xuehao)),
                                   teacher_id=Teachers.objects.get(teacher_id=str(teacher_id)))

        msg = "成功上传教师表！！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    except Exception as e:
        print(e)
    return HttpResponse('请返回重新提交表格!！')


# ———————题库—————————
def myadmin_tiku(request):
    tiku_list = TiKu_1.objects.filter().order_by('id')  # 班级过滤器

    context = {"tiku_list": tiku_list}
    return render(request, 'myadmin/tiku/myadmin_tiku.html', context)


# 添加题库
def myadmin_tiku_add(request):
    if request.method == "GET":
        tiku = TiKu_1.objects.filter(is_active=True).order_by()

        if tiku:
            id = tiku.last().id
            if tiku.count() == 10:
                msg = '只能添加10道题！'
                return render(request, 'myadmin/tishi.html', {"msg": msg})
            else:
                return render(request, 'myadmin/tiku/tiku_add.html', {"id": id + 1})
        else:
            id = 1

        return render(request, 'myadmin/tiku/tiku_add.html', {"id": id})
    if request.method == "POST":
        id = request.POST['id']
        timu = request.POST['timu']

        try:
            stu = TiKu_1.objects.filter(str(id))
        except Exception as e:
            print(e)

        stu_add = TiKu_1.objects.create(id=str(id), timu=timu)
        msg = '添加成功！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})


# 编辑题库
def myadmin_tiku_edit(request, id):
    try:
        tiku = TiKu_1.objects.get(id=int(id))


    except Exception as e:
        print(e)
        msg = '没有此评价题！！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    tiku = TiKu_1.objects.get(id=id)
    if request.method == 'GET':

        return render(request, 'myadmin/tiku/tiku_edit.html', {"tiku": tiku})
    else:

        timu = request.POST['timu']

        # is_active = request.POST['is_active']

        tiku.timu = timu

        # stu.is_active = is_active
        tiku.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})


# 删除题库
def myadmin_tiku_del(request, id):
    try:
        tiku = TiKu_1.objects.get(id=int(id))


    except Exception as e:
        print(e)
        msg = '没有此评价题！！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    tiku.delete()
    msg = '删除成功！'
    return render(request, 'myadmin/tishi.html', {"msg": msg})


# 管理员评价成绩查看模块
def myadmin_pingjia(request):
    from django.db.models import Avg
    teachers_list = Teachers.objects.filter(is_active=True)
    # 评价
    data_list = list()
    data = list()
    pingjia = PingJia.objects
    for i in teachers_list:
        my_list = list()
        my_list.append(i.teacher_id)
        my_list.append(i.name)
        if i.sex == "male":
            sex = "男"
        elif i.sex == "female":
            sex = "女"
        my_list.append(sex)
        my_list.append(i.phone)
        # 学生评价率
        pingjia_ok = pingjia.filter(kecheng__teacher_id=i.teacher_id, is_active=True).count()
        pingjia_sum = KeCheng.objects.filter(is_active=True, teacher_id=i.teacher_id).count()
        # print(pingjia_sum, "---", pingjia_ok)
        try:
            pjl = pingjia_ok / pingjia_sum
            pjl = str(float('%.2f' % pjl)) + '%'
            my_list.append(pjl)
        except Exception as e:
            print("myadmin_pingjia-error:", e)
            my_list.append('未授课')
        avg = pingjia.filter(kecheng__teacher_id=i.teacher_id, kecheng__ok='ok').aggregate(Avg('s_avg'))
        # print(avg)
        avg = avg['s_avg__avg']
        try:
            pjl = str(float('%.2f' % avg)) + '%'
            my_list.append(pjl)
        except Exception as e:
            print(e)
            my_list.append('未评价')

        data.append(my_list)

    return render(request, 'myadmin/pingjia/myadmin_pingjia.html', {"data_list": data_list, "data": data})


def myadmin_pingjia_pjl(request):
    """
    1.过滤器显示班级
    2.学生数量
    3.已评价
    4.未评价
    5.评价率
    :param request:
    :return:
    """
    data_list = list()
    banji_data = Students.objects.filter(is_active=True).values('banji').distinct()

    # banji_data = banji_data.distinct('banji')
    for i in banji_data:
        list1 = list()
        banji = i['banji']
        list1.append(banji)
        stu_sum = Students.objects.filter(is_active=True, banji=banji).count()
        list1.append(stu_sum)
        ok_pingjia = KeCheng.objects.filter(ok='ok', xuehao__banji=banji).values("xuehao").distinct().count()
        list1.append(ok_pingjia)

        try:
            no_pingjia = stu_sum - ok_pingjia
            list1.append(no_pingjia)
            pjl = ok_pingjia / stu_sum
            pjl = str(float('%.2f' % pjl)) + '%'
            list1.append(pjl)
        except Exception as e:
            print('myadmin_pingjia_PJL%s' % e)
        data_list.append(list1)
    return render(request, 'myadmin/pingjia/myadmin_pingjia_pjl.html', {"data_list": data_list})


def myadmin_pingjia_not(request):
    try:
        not_pingjia = KeCheng.objects.filter(is_active=True, ok="no").order_by("xuehao").distinct().values(
            "xuehao",
            "xuehao__name",
            "xuehao__banji",
            "xuehao__phone",
            "xuehao__email")
        data_list = list()
        for i in not_pingjia:
            list1 = list()
            xuehao = i['xuehao']
            name = i['xuehao__name']
            banji = i['xuehao__banji']
            phone = i['xuehao__phone']
            email = i['xuehao__email']
            sum = KeCheng.objects.filter(xuehao=xuehao, ok="no").count()
            list1.append(xuehao)
            list1.append(name)
            list1.append(banji)
            list1.append(phone)
            list1.append(email)
            list1.append(sum)
            data_list.append(list1)
    except Exception as e:
        print("myadmin_pingjia_not：", e)
    return render(request, 'myadmin/pingjia/myadmin_pingjia_not.html', {"data_list": data_list})


def myadmin_admin(request):
    data_list = GuanLiYuan.objects.filter(is_active=True).values("name", "phone", "email", "zhiwu")

    return render(request, 'myadmin/admin/myadmin_admin.html', {"data_list": data_list})


def myadmin_admin_add(request):
    if request.method == "GET":
        return render(request, 'myadmin/admin/admin_add.html')
    if request.method == "POST":
        name = request.POST['name']
        phone = request.POST['phone']
        email = request.POST['email']
        zhiwu = request.POST['zhiwu']

        try:
            stu = GuanLiYuan.objects.filter(name=name)
        except Exception as e:
            print(e)
        if stu:
            msg = '用户已存在！'
            return render(request, 'myadmin/admin/admin_add.html', {"msg": msg})
        else:
            stu_add = GuanLiYuan.objects.create(name=name,
                                                phone=phone,
                                                email=email,
                                                zhiwu=zhiwu)
            msg = '添加成功！'
            return render(request, 'myadmin/tishi.html', {"msg": msg})


def myadmin_admin_edit(request, name):
    try:
        stu = GuanLiYuan.objects.get(name=name)
    except Exception as e:
        print(e)
        msg = '没有此管理员！'
        return render(request, 'myadmin/tishi.html', {"msg": msg})

    if request.method == 'GET':

        return render(request, 'myadmin/admin/admin_edit.html', locals())
    else:

        phone = request.POST['phone']
        email = request.POST['email']
        zhiwu = request.POST['email']

        stu.name = name
        stu.phone = phone

        stu.email = email
        stu.zhiwu = zhiwu
        stu.save()
        msg = '修改成功！'
        return render(request, "myadmin/tishi.html", {"msg": msg})


def myadmin_admin_del(request, name):
    try:
        stu = GuanLiYuan.objects.get(name=name)
    except Exception as e:
        print(e)
        msg = "没有此管理员！"
        return render(request, 'myadmin/tishi.html', {"msg": msg})
    stu.is_active = False
    stu.save()
    msg = "删除成功！"
    return render(request, 'myadmin/tishi.html', {"msg": msg})


def pswd_update(request):
    if request.method == 'GET':
        return render(request, 'myadmin/pswd_updat.html')
    if request.method == 'POST':
        name = request.session.get('name')
        pswd = request.POST['pswd']
        pswd_1 = request.POST['pswd_1']
        pswd_2 = request.POST['pswd_2']

        # 哈希算法
        m = hashlib.md5()
        m.update(pswd.encode())
        password_m = m.hexdigest()

        if pswd_1 != pswd_2:
            msg = "新密码不一致！！！"
            return render(request, 'myadmin/pswd_updat.html', {"msg": msg})
        try:
            ss = GuanLiYuan.objects.filter(name=name, password=password_m, is_active=True)
            sm = GuanLiYuan.objects.get(name=name, is_active=True)
        except Exception as e:
            print("myadmin_update_pswd:", e)
        if ss:
            # 哈希算法
            m = hashlib.md5()
            m.update(pswd_1.encode())
            password_m = m.hexdigest()
            sm.password = password_m
            sm.save()
            msg = "密码修改成功！！"
            return render(request, 'index.html', {"msg": msg})
        else:
            msg = "原密码错误！"
            return render(request, 'myadmin/pswd_updat.html', {"msg": msg})


def myadmin_pingjia_show(request):

    from django.db.models import Avg
    teacher_id = request.GET.get('id')
    teacher=Teachers.objects.get(teacher_id=teacher_id)
    pingjiabaio = PingJia.objects.filter(is_active=True,
                                         # kecheng__xuehao__banji=banji_id,
                                         kecheng__teacher_id=teacher_id)  # 评价
    tiku = TiKu_1.objects.filter(is_active=True)  # 题库
    a = KeCheng.objects.filter(teacher_id=teacher_id, is_active=True) \
        .values('xuehao__banji', 'xuehao', 'id', 'ok')  # 班级过滤器
    b = KeCheng.objects.filter(teacher_id=teacher_id, is_active=True, )

    # 评价率  评价率 = 学生评价数  /  学生数
    pingjia_sum = KeCheng.objects.filter(is_active=True, ok='ok').count()  # 学生评价书
    stu_sum = Students.objects.filter(is_active=True).count()
    PJL = float('%.2f' % (pingjia_sum / stu_sum * 100))

    # 求平均值
    avg = PingJia.objects.filter(kecheng__ok='ok',
                                 kecheng__teacher_id=teacher_id,
                                 # kecheng__xuehao__banji=banji_id,
                                 ).aggregate(Avg("s_daan1"),
                                             Avg("s_daan2"),
                                             Avg("s_daan3"),
                                             Avg("s_daan4"),
                                             Avg("s_daan5"),
                                             Avg("s_daan6"),
                                             Avg("s_daan7"),
                                             Avg("s_daan8"),
                                             Avg("s_daan9"),
                                             Avg("s_daan10"))
    print(avg)

    try:
        avg1 = float('%.2f' % avg['s_daan1__avg'])
        avg2 = float('%.2f' % avg['s_daan2__avg'])
        avg3 = float('%.2f' % avg['s_daan3__avg'])
        avg4 = float('%.2f' % avg['s_daan4__avg'])
        avg5 = float('%.2f' % avg['s_daan5__avg'])
        avg6 = float('%.2f' % avg['s_daan6__avg'])
        avg7 = float('%.2f' % avg['s_daan7__avg'])
        avg8 = float('%.2f' % avg['s_daan8__avg'])
        avg9 = float('%.2f' % avg['s_daan9__avg'])
        avg10 = float('%.2f' % avg['s_daan10__avg'])
        # 综合评价
        s = 0
        s2 = 0
        for i in avg:
            if avg[i] != 0:
                s += 1
                s2 += avg[i]
        s_avg = s2 / s
        s_avg = float('%.2f' % s_avg)
    except Exception as e:
        print('ok')
    return render(request,'myadmin/pingjia/pingjia_show.html',locals())